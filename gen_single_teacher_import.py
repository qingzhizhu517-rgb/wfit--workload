# -*- coding: utf-8 -*-
"""
生成「单个教师·多门课程」教学任务导入测试 Excel。
严格对齐 TeachingTaskImportDTO(15 列表头顺序) + validateRow 校验规则。

目标教师：陶凤敏 / 工号 T20270005 / 教授 / 专任 / 智能制造学院
选此人原因：教授单价 70 元且性质为专任(才计绩效酬金)，可完整跑通
「导入 → 汇总重算 → 酬金结算 → 报表导出 → 教师确认 → 签字」全流程。

覆盖维度：
  - 工作量类别 G1~G6 全覆盖
  - 授课层次 本科/专科
  - 专业大类 理工类/文史类/艺术类
  - 课程性质 必修/选修 (影响 K1: 必修1.1 选修1.0)
  - 课程级别×角色 省级一流·主持人 / 校级精品·团队前3 / 其他·独立 (影响 Q2)
  - 教学评价 优秀/良好/合格/不合格 (不合格 → Q1=0.8)
  - 合堂人数 常规 / ≥120 / ≥151 (影响 N: 1.0/1.1/1.2)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADERS = ["学年学期", "教师工号", "教师姓名", "课程名称", "课程代码", "工作量类别",
           "授课层次", "专业大类", "课程性质", "课程级别", "课程角色", "教学评价",
           "选课人数", "计划学时/天数/周数", "课程系数"]

SEM = "2025-2026-1"
GH = "T20270005"
XM = "陶凤敏"

# 每行： (课程名, 代码, 类别, 层次, 大类, 性质, 级别, 角色, 评价, 人数, 计划值, 系数)
# 计划值：G1/G2=学时  G3=天数  G4=学分  G5=占位(>0,计算用人数×系数)  G6=周数
# 系数：G2=K(专业) G3=D(指导) G5=K5(单位)；G1/G4/G6 留空(引擎按规则/默认取)
ROWS = [
    # G1 理论课 —— 覆盖 省级一流·主持人 + 大合堂(130人→N=1.1) + 必修
    ("高等数学",   "MATH1001", "G1", "本科", "理工类", "必修", "省级一流", "主持人",  "优秀",   130, 64, ""),
    # G1 —— 不合格(Q1=0.8) + 超大合堂(155人→N=1.2) + 必修
    ("数据结构",   "CS2001",   "G1", "本科", "理工类", "必修", "其他",     "独立",    "不合格", 155, 48, ""),
    # G1 —— 选修(K1=1.0) + 校级精品·团队前3(Q2=1.1) + 文史类
    ("大学语文",   "CHN1002",  "G1", "本科", "文史类", "选修", "校级精品", "团队前3", "良好",   90,  32, ""),
    # G1 —— 专科 + 艺术类 + 选修 + 常规
    ("艺术鉴赏",   "ART1003",  "G1", "专科", "艺术类", "选修", "其他",     "独立",    "合格",   45,  32, ""),
    # G2 课内实践 —— 系数K=1.0(理工)
    ("软件工程实训", "SE3001",  "G2", "本科", "理工类", "必修", "其他",     "独立",    "优秀",   40,  32, 1.0),
    # G2 —— 选修
    ("数据库实践",  "DB3002",   "G2", "本科", "理工类", "选修", "其他",     "独立",    "良好",   35,  24, 1.0),
    # G3 实习实训 —— 天数2天(×8学时) 指导系数D=4.0(理工)
    ("生产实习",   "PRAC4001", "G3", "本科", "理工类", "必修", "其他",     "独立",    "优秀",   30,  2,  4.0),
    # G4 课程设计 —— 学分2 指导人数35(≤60)
    ("综合课程设计", "CD5001",  "G4", "本科", "理工类", "必修", "其他",     "独立",    "优秀",   35,  2,  ""),
    # G5 毕业论文 —— 指导5人 单位系数K5=9(理工本科)
    ("毕业设计",   "GRAD6001", "G5", "本科", "理工类", "必修", "其他",     "独立",    "优秀",   5,   1,  9),
    # G6 集中实习 —— 周数4 指导20人(≤20)
    ("顶岗实习",   "INT7001",  "G6", "本科", "理工类", "必修", "其他",     "独立",    "良好",   20,  4,  ""),
]

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "教学任务导入"

    head_fill = PatternFill("solid", fgColor="4472C4")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    for r in ROWS:
        name, code, gt, edu, major, nature, level, role, ev, cnt, base, coef = r
        ws.append([SEM, GH, XM, name, code, gt, edu, major, nature, level, role, ev, cnt, base, coef])

    for row in ws.iter_rows(min_row=2, max_row=1 + len(ROWS), max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = center
            cell.border = border

    widths = [14, 12, 10, 18, 12, 11, 10, 10, 10, 12, 11, 10, 10, 18, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    ws.freeze_panes = "A2"

    out = "教学任务导入_陶凤敏_10门课_全流程测试.xlsx"
    wb.save(out)
    print("已生成:", out, "共", len(ROWS), "行")

if __name__ == "__main__":
    build()
