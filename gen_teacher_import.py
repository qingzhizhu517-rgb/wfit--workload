# -*- coding: utf-8 -*-
"""生成教师档案批量导入测试 Excel（表头/取值严格对齐后端校验）。"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 表头顺序必须与 TeacherProfileImportDTO 的 @ExcelProperty 顺序完全一致
HEADERS = ["教师工号", "教师姓名", "院部名称", "职称", "人员性质", "手机号", "邮箱"]

# 只能用库中真实存在的部门（findDeptId 找不到会整行失败）
DEPTS = ["潍坊理工学院", "智能制造学院", "大数据学院", "信管本"]
# 校验白名单
TITLES = ["教授", "副教授", "讲师", "助教", "未定级"]
NATURES = ["专任", "外聘", "校企", "银龄", "青州外聘"]

SURNAMES = list("赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜")
GIVEN = list("伟芳娜秀英敏静丽强磊军洋勇艳杰娟涛明超霞平刚桂兰凤云鹏浩宇泽")

def rand_name():
    n = random.choice([2, 3])
    return random.choice(SURNAMES) + "".join(random.choice(GIVEN) for _ in range(n - 1))

def rand_phone():
    return "1" + random.choice("3567889") + "".join(random.choice("0123456789") for _ in range(9))

def build_rows(n):
    rows = []
    for i in range(1, n + 1):
        code = f"T2027{i:04d}"          # 全新工号，避开库中已有账号
        name = rand_name()
        dept = DEPTS[i % len(DEPTS)]
        title = TITLES[i % len(TITLES)]
        nature = NATURES[i % len(NATURES)]
        phone = rand_phone() if i % 5 != 0 else ""   # 每 5 行留 1 个空手机号（可选字段）
        email = f"{code.lower()}@wfit.edu.cn" if i % 4 != 0 else ""
        rows.append([code, name, dept, title, nature, phone, email])
    return rows

def write_book(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "教师档案"
    head_fill = PatternFill("solid", fgColor="409EFF")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    widths = [14, 12, 20, 12, 12, 14, 24]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return len(rows)

if __name__ == "__main__":
    random.seed(20270826)
    base = r"C:\Users\Aohs\Desktop"

    # 1) 正常数据：30 行，全部应导入成功
    ok = build_rows(30)
    n1 = write_book(base + r"\教师批量导入_测试_30条.xlsx", ok)

    # 2) 含错误行：在正常数据基础上追加 5 个典型错误，测试逐行校验与失败明细
    bad = build_rows(10)
    bad.append(["T2027E001", "错误一", "不存在的学院", "讲师", "专任", "13800001111", "e1@wfit.edu.cn"])  # 部门不存在
    bad.append(["T2027E002", "错误二", "大数据学院", "研究员", "专任", "13800002222", "e2@wfit.edu.cn"])   # 职称非法
    bad.append(["T2027E003", "错误三", "信管本", "讲师", "临时工", "13800003333", "e3@wfit.edu.cn"])       # 人员性质非法
    bad.append(["", "错误四", "信管本", "讲师", "专任", "13800004444", "e4@wfit.edu.cn"])                   # 工号为空
    bad.append(["T2027E005", "错误五", "信管本", "讲师", "专任", "123", "bad-email"])                        # 手机号+邮箱格式错
    n2 = write_book(base + r"\教师批量导入_含错误行_15条.xlsx", bad)

    print(f"生成完成：\n  正常 {n1} 行 -> 教师批量导入_测试_30条.xlsx\n  含错误 {n2} 行 -> 教师批量导入_含错误行_15条.xlsx")
