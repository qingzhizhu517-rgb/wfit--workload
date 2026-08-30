# -*- coding: utf-8 -*-
"""生成教学任务批量导入测试 Excel（严格对齐 TeachingTaskImportDTO + validateRow）。"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 表头顺序必须与 TeachingTaskImportDTO 的 @ExcelProperty 顺序完全一致（15 列）
HEADERS = ["学年学期", "教师工号", "教师姓名", "课程名称", "课程代码", "工作量类别",
           "授课层次", "专业大类", "课程性质", "课程级别", "课程角色", "教学评价",
           "选课人数", "计划学时/天数/周数", "课程系数"]

SEMESTER = "2025-2026-1"          # 合法学期（秋季学期号=1）
# 现有 30 位教师工号 T20270001~T20270030（导入按工号精确查找，必须已存在）
TEACHERS = [(f"T2027{i:04d}", n) for i, n in enumerate(
    ["钱伟","沈涛","王杰鹏","孔磊娜","陶凤敏","严泽","韩英勇","何凤","朱丽","孔霞刚",
     "张秀","华艳","赵军涛","严磊桂","孔娟","李明静","郑凤鹏","许娜","许磊","施强",
     "何勇艳","赵鹏艳","李云泽","孔军浩","施英","郑强","张浩","何芳平","金明超","钱浩"], start=1)]

EDU = ["本科", "专科"]
MAJOR = ["理工类", "文史类", "艺术类", "其他"]
NATURE = ["必修", "选修"]
LEVEL = ["省级一流", "校级精品", "其他"]
ROLE = ["主持人", "团队前3", "独立"]
EVAL = ["优秀", "良好", "合格"]

# 每种类别的"计划学时/天数/周数"合理取值 + 课程系数
COURSE_NAMES = {
    "G1": ["高等数学", "大学英语", "程序设计基础", "线性代数", "数据结构", "操作系统", "计算机网络", "概率论"],
    "G2": ["电工实验", "物理实验", "化学实验", "软件工程实训", "数据库实践"],
    "G3": ["认识实习", "生产实习", "专业实训", "工程实践"],
    "G4": ["课程设计", "综合课程设计", "系统课程设计"],
    "G5": ["毕业论文", "毕业设计"],
    "G6": ["集中实习", "顶岗实习", "毕业实习"],
}

def gen_row(i, teacher):
    code, name = teacher
    # 让 6 类都覆盖到：前 6 行强制 G1~G6，其余随机偏向 G1/G2
    if i <= 6:
        wtype = f"G{i}"
    else:
        wtype = random.choices(["G1","G2","G3","G4","G5","G6"], weights=[5,3,1,1,1,1])[0]
    course = random.choice(COURSE_NAMES[wtype])
    ccode = f"{wtype}{random.randint(1000,9999)}"
    edu = random.choice(EDU)
    major = random.choice(MAJOR)
    nature = random.choice(NATURE)
    level = random.choices(LEVEL, weights=[1,1,6])[0]
    role = random.choice(ROLE)
    ev = random.choices(EVAL, weights=[2,3,1])[0]

    if wtype == "G1":       # 理论课：计划学时 32~64，学生数影响合堂系数
        base, students, coef = random.choice([32,48,64]), random.randint(30,160), ""
    elif wtype == "G2":     # 实践课：实践学时；课程系数=K(理工1.0/其他0.9)
        base, students, coef = random.choice([16,24,32]), random.randint(20,60), (1.0 if major=="理工类" else 0.9)
    elif wtype == "G3":     # 实习实训：天数；系数=D(理工4/艺术3/文史2)
        base, students, coef = random.choice([5,10,15]), random.randint(20,50), (4.0 if major=="理工类" else 3.0 if major=="艺术类" else 2.0)
    elif wtype == "G4":     # 课程设计：学分（base）；R4=选课人数（上限60）
        base, students, coef = random.choice([1,2,3]), random.randint(20,80), ""
    elif wtype == "G5":     # 毕业论文：R5=人数；系数=K5（理工本9/专5）
        base, students, coef = 1, random.randint(3,12), (9 if edu=="本科" else 5)
    else:                   # G6 集中实习：周数（base）；R6=人数（上限20）
        base, students, coef = random.choice([2,4,6]), random.randint(10,30), ""

    return [SEMESTER, code, name, course, ccode, wtype, edu, major, nature, level, role, ev,
            students, base, coef]

def write_book(path, rows):
    wb = Workbook(); ws = wb.active; ws.title = "教学任务"
    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor="409EFF"); font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    widths = [18,12,10,24,14,12,10,10,10,12,12,10,10,18,10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(path); return len(rows)

if __name__ == "__main__":
    random.seed(20270826)
    base = r"C:\Users\Aohs\Desktop"

    # 正常数据：每位教师 1 门课 = 30 行，G1~G6 全覆盖
    rows = [gen_row(i, t) for i, t in enumerate(TEACHERS, start=1)]
    n1 = write_book(base + r"\教学任务导入_测试_30条.xlsx", rows)

    # 含错误行：10 正常 + 5 典型错误，测逐行校验/失败明细
    bad = [gen_row(i, t) for i, t in enumerate(TEACHERS[:10], start=1)]
    bad.append(["2025-1", "T20270001", "错误一", "测试课", "X1", "G1", "本科", "理工类", "必修", "其他", "独立", "良好", 30, 48, ""])      # 学期格式错
    bad.append([SEMESTER, "T99999999", "错误二", "测试课", "X2", "G1", "本科", "理工类", "必修", "其他", "独立", "良好", 30, 48, ""])       # 工号不存在
    bad.append([SEMESTER, "T20270002", "错误三", "测试课", "X3", "G9", "本科", "理工类", "必修", "其他", "独立", "良好", 30, 48, ""])       # 类别非 G1-G6
    bad.append([SEMESTER, "T20270003", "错误四", "", "X4", "G1", "本科", "理工类", "必修", "其他", "独立", "良好", 30, 48, ""])            # 课程名为空
    bad.append([SEMESTER, "T20270004", "错误五", "测试课", "X5", "G1", "本科", "理工类", "必修", "其他", "独立", "良好", 30, 0, ""])        # 计划学时=0
    n2 = write_book(base + r"\教学任务导入_含错误行_15条.xlsx", bad)

    print(f"OK: normal={n1} rows, with-errors={n2} rows")
